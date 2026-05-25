#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Dispatch calendar Excel-to-TSV converter (CMD version)."""

from __future__ import annotations

import sys
import traceback
import re
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def write_error_text(psz_excel_file_path: str, psz_error_message: str) -> str:
    """Write an error text file beside the source Excel file."""
    obj_excel_path: Path = Path(psz_excel_file_path)
    psz_error_file_name: str = f"{obj_excel_path.stem}_error.txt"
    obj_error_file_path: Path = obj_excel_path.with_name(psz_error_file_name)

    with obj_error_file_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_error_file:
        obj_error_file.write(psz_error_message)

    return str(obj_error_file_path)


def normalize_line_breaks_and_trim(psz_text: str) -> str:
    """Normalize line breaks and trim spaces for strict matching."""
    psz_normalized_text: str = psz_text.replace("\r\n", "\n").strip()
    return psz_normalized_text


def should_skip_row_by_first_column(list_source_row: list[str], list_skip_keywords: list[str]) -> bool:
    """Return True when first-column text matches any normalized skip keyword exactly."""
    if len(list_source_row) == 0:
        return False

    psz_first_column_text: str = normalize_line_breaks_and_trim(list_source_row[0])
    if psz_first_column_text == "":
        return False

    for psz_skip_keyword in list_skip_keywords:
        psz_normalized_skip_keyword: str = normalize_line_breaks_and_trim(psz_skip_keyword)
        if psz_first_column_text == psz_normalized_skip_keyword:
            return True

    return False




def is_shifted_continuation_row(list_previous_row: list[str], list_current_row: list[str]) -> bool:
    """Detect left-shifted continuation rows that should map A/B into C/D."""
    if len(list_previous_row) < 4 or len(list_current_row) < 2:
        return False

    psz_previous_name: str = normalize_line_breaks_and_trim(list_previous_row[0])
    psz_previous_car_number: str = normalize_line_breaks_and_trim(list_previous_row[1])
    psz_previous_spare_car: str = normalize_line_breaks_and_trim(list_previous_row[2])
    psz_current_first_value: str = normalize_line_breaks_and_trim(list_current_row[0])
    psz_current_second_value: str = normalize_line_breaks_and_trim(list_current_row[1])

    if psz_previous_name == "" or psz_previous_car_number == "" or psz_previous_spare_car == "":
        return False

    if psz_current_first_value == "" or psz_current_second_value == "":
        return False

    for psz_remaining_value in list_current_row[2:]:
        if normalize_line_breaks_and_trim(psz_remaining_value) != "":
            return False

    return True


def shift_row_to_the_right(list_current_row: list[str], iShiftCount: int) -> list[str]:
    """Return a row shifted right by adding empty cells at the beginning."""
    return [""] * iShiftCount + list_current_row
def merge_continuation_rows(list_source_rows: list[list[str]]) -> list[list[str]]:
    """Merge continuation rows into previous row by joining values with newlines per column."""
    list_merged_rows: list[list[str]] = []

    for list_current_row in list_source_rows:
        b_has_previous_row: bool = len(list_merged_rows) > 0
        b_is_blank_first_column_continuation: bool = b_has_previous_row and normalize_line_breaks_and_trim(list_current_row[0]) == ""

        b_has_shift_adjustment: bool = b_has_previous_row and is_shifted_continuation_row(list_merged_rows[-1], list_current_row)
        if b_has_shift_adjustment:
            list_adjusted_current_row: list[str] = shift_row_to_the_right(list_current_row, 2)
        else:
            list_adjusted_current_row = list_current_row

        b_is_continuation_row: bool = b_has_previous_row and (b_is_blank_first_column_continuation or b_has_shift_adjustment)

        if not b_is_continuation_row:
            list_merged_rows.append(list_current_row[:])
            continue

        list_previous_row: list[str] = list_merged_rows[-1]
        i_previous_column_count: int = len(list_previous_row)
        i_current_column_count: int = len(list_adjusted_current_row)
        i_max_column_count: int = i_previous_column_count if i_previous_column_count > i_current_column_count else i_current_column_count

        if i_previous_column_count < i_max_column_count:
            list_previous_row.extend([""] * (i_max_column_count - i_previous_column_count))

        for iColumnIndex in range(i_max_column_count):
            psz_current_value: str = ""
            if iColumnIndex < i_current_column_count:
                psz_current_value = list_adjusted_current_row[iColumnIndex]

            if normalize_line_breaks_and_trim(psz_current_value) == "":
                continue

            psz_previous_value: str = list_previous_row[iColumnIndex]
            if normalize_line_breaks_and_trim(psz_previous_value) == "":
                list_previous_row[iColumnIndex] = psz_current_value
            else:
                list_previous_row[iColumnIndex] = f"{psz_previous_value}\n{psz_current_value}"

    return list_merged_rows




def escape_newlines_in_cell_text(psz_cell_text: str) -> str:
    """Normalize embedded newlines in cell text."""
    psz_normalized_text: str = psz_cell_text.replace("\r\n", "\n").replace("\r", "\n")
    return psz_normalized_text


def expand_rows_by_embedded_newlines(
    list_source_rows: list[list[str]],
    i_fixed_column_count: int = 3,
) -> list[list[str]]:
    """Expand embedded newlines into physical TSV rows while blanking fixed leading columns on continuation lines."""
    list_expanded_rows: list[list[str]] = []

    for list_source_row in list_source_rows:
        list_split_cells: list[list[str]] = []
        i_max_split_count: int = 1
        b_has_multiline_non_fixed_column: bool = False

        for i_column_index, psz_cell_text in enumerate(list_source_row):
            list_cell_lines: list[str] = psz_cell_text.split("\n")
            list_split_cells.append(list_cell_lines)
            if len(list_cell_lines) > i_max_split_count:
                i_max_split_count = len(list_cell_lines)
            if i_column_index >= i_fixed_column_count and len(list_cell_lines) > 1:
                b_has_multiline_non_fixed_column = True

        for i_row_index in range(i_max_split_count):
            list_expanded_row: list[str] = []

            for i_column_index, list_cell_lines in enumerate(list_split_cells):
                psz_cell_value: str = list_cell_lines[i_row_index] if i_row_index < len(list_cell_lines) else ""

                b_should_blank_fixed_column: bool = (
                    i_row_index > 0
                    and b_has_multiline_non_fixed_column
                    and i_column_index < i_fixed_column_count
                    and len(list_cell_lines) == 1
                )
                if b_should_blank_fixed_column:
                    psz_cell_value = ""

                list_expanded_row.append(psz_cell_value)

            list_expanded_rows.append(list_expanded_row)

    return list_expanded_rows


def normalize_file_stem_for_step_output(psz_file_stem: str) -> str:
    """Replace half/full-width spaces in file stem with underscore."""
    return re.sub(r"[ \u3000]+", "_", psz_file_stem)


def create_step0001_tsv_from_tsv(psz_tsv_file_path: str) -> str:
    """Create step0001 TSV by removing spare-car column from the generated TSV."""
    obj_tsv_file_path: Path = Path(psz_tsv_file_path)
    psz_normalized_stem: str = normalize_file_stem_for_step_output(obj_tsv_file_path.stem)
    obj_step_tsv_path: Path = obj_tsv_file_path.with_name(f"{psz_normalized_stem}_step0001.tsv")

    list_output_lines: list[str] = []
    with obj_tsv_file_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        for psz_line in obj_input_file:
            psz_line_without_newline: str = psz_line.rstrip("\r\n")
            list_columns: list[str] = psz_line_without_newline.split("\t")
            if len(list_columns) >= 3:
                del list_columns[2]
            list_output_lines.append("\t".join(list_columns))

    with obj_step_tsv_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_output_file:
        for psz_output_line in list_output_lines:
            obj_output_file.write(psz_output_line + "\n")

    return str(obj_step_tsv_path)


def parse_step0001_tsv_to_calendar_records(psz_step0001_tsv_path: str) -> list[dict[str, Any]]:
    """Parse step0001 TSV into per-person calendar records."""
    obj_step0001_tsv_path: Path = Path(psz_step0001_tsv_path)
    with obj_step0001_tsv_path.open(mode="r", encoding="utf-8", newline="") as obj_input_file:
        list_lines: list[str] = [psz_line.rstrip("\r\n") for psz_line in obj_input_file]

    if len(list_lines) < 3:
        return []

    psz_work_date_text: str = list_lines[0]
    list_data_lines: list[str] = list_lines[2:]

    list_records: list[dict[str, Any]] = []
    obj_current_record: dict[str, Any] | None = None

    for psz_line in list_data_lines:
        list_columns: list[str] = psz_line.split("\t")
        if len(list_columns) < 8:
            list_columns.extend([""] * (8 - len(list_columns)))

        psz_name: str = list_columns[0].strip()
        psz_car_no: str = list_columns[1].strip()
        list_slot_values: list[str] = [list_columns[i].strip() for i in range(2, 8)]

        b_is_blank_row: bool = psz_name == "" and psz_car_no == "" and all(psz_slot == "" for psz_slot in list_slot_values)
        if b_is_blank_row:
            continue

        if psz_name != "":
            if obj_current_record is not None:
                list_records.append(obj_current_record)
            obj_current_record = {
                "name": psz_name,
                "car_nos": [],
                "slots": {str(i): "" for i in range(1, 7)},
                "work_date_text": psz_work_date_text,
            }

        if obj_current_record is None:
            continue

        if psz_car_no != "":
            list_car_nos: list[str] = obj_current_record["car_nos"]
            if psz_car_no not in list_car_nos and len(list_car_nos) < 3:
                list_car_nos.append(psz_car_no)

        for i_slot_index, psz_slot_value in enumerate(list_slot_values, start=1):
            if psz_slot_value == "":
                continue
            psz_slot_key: str = str(i_slot_index)
            psz_existing: str = obj_current_record["slots"][psz_slot_key]
            if psz_existing == "":
                obj_current_record["slots"][psz_slot_key] = psz_slot_value
            else:
                obj_current_record["slots"][psz_slot_key] = f"{psz_existing}\n{psz_slot_value}"

    if obj_current_record is not None:
        list_records.append(obj_current_record)

    for obj_record in list_records:
        list_car_nos = obj_record["car_nos"]
        obj_record["car_no"] = ",".join(list_car_nos)
        obj_record["car_no_display"] = "/".join(list_car_nos)
        psz_car_for_title: str = obj_record["car_no_display"] if obj_record["car_no_display"] != "" else "車番未設定"
        list_slot_lines: list[list[str]] = [obj_record["slots"][str(i_slot_index)].split("\n") for i_slot_index in range(1, 7)]
        i_max_slot_line_count: int = 1
        for list_slot_line in list_slot_lines:
            if len(list_slot_line) > i_max_slot_line_count:
                i_max_slot_line_count = len(list_slot_line)

        def build_slot_line(i_line_index: int) -> str:
            list_values: list[str] = [
                list_slot_line[i_line_index] if i_line_index < len(list_slot_line) else "" for list_slot_line in list_slot_lines
            ]
            list_non_blank_values: list[str] = [psz_value for psz_value in list_values if psz_value != ""]
            return " ".join(list_non_blank_values)

        psz_first_slot_line: str = build_slot_line(0)
        list_title_parts: list[str] = [obj_record["name"], psz_car_for_title]
        if psz_first_slot_line != "":
            list_title_parts.append(psz_first_slot_line)
        obj_record["title_text"] = " ".join(list_title_parts)

        list_body_lines: list[str] = []
        if i_max_slot_line_count >= 2:
            psz_second_slot_line: str = build_slot_line(1)
            if psz_second_slot_line != "":
                list_body_lines.append(psz_second_slot_line)
        else:
            if psz_first_slot_line != "":
                list_body_lines.append(psz_first_slot_line)

        for i_line_index in range(2, i_max_slot_line_count):
            psz_extra_slot_line: str = build_slot_line(i_line_index)
            if psz_extra_slot_line != "":
                list_body_lines.append(psz_extra_slot_line)

        obj_record["body_text"] = "\n".join(list_body_lines)

    return list_records


def create_step0002_outputs_from_step0001_tsv(psz_step0001_tsv_path: str) -> tuple[str, str]:
    """Create step0002 TSV and JSON (NDJSON) from step0001 TSV."""
    obj_step0001_tsv_path: Path = Path(psz_step0001_tsv_path)
    psz_step_stem: str = obj_step0001_tsv_path.stem
    psz_output_stem: str = psz_step_stem[:-9] if psz_step_stem.endswith("_step0001") else psz_step_stem

    obj_step0002_tsv_path: Path = obj_step0001_tsv_path.with_name(f"{psz_output_stem}_step0002.tsv")
    obj_step0002_json_path: Path = obj_step0001_tsv_path.with_name(f"{psz_output_stem}_step0002.json")

    list_records: list[dict[str, Any]] = parse_step0001_tsv_to_calendar_records(str(obj_step0001_tsv_path))

    list_tsv_columns: list[str] = [
        "name",
        "car_no_display",
        "car_no",
        "car_nos_joined",
        "slot1",
        "slot2",
        "slot3",
        "slot4",
        "slot5",
        "slot6",
        "title_text",
        "body_text",
        "work_date_text",
    ]
    with obj_step0002_tsv_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_tsv_file:
        obj_tsv_file.write("\t".join(list_tsv_columns) + "\n")
        for obj_record in list_records:
            list_slot_lines: list[list[str]] = [obj_record["slots"][str(i_slot_index)].split("\n") for i_slot_index in range(1, 7)]
            i_max_slot_line_count: int = 1
            for list_slot_line in list_slot_lines:
                if len(list_slot_line) > i_max_slot_line_count:
                    i_max_slot_line_count = len(list_slot_line)

            for i_line_index in range(i_max_slot_line_count):
                list_slot_values: list[str] = [
                    list_slot_line[i_line_index] if i_line_index < len(list_slot_line) else "" for list_slot_line in list_slot_lines
                ]
                list_row_values: list[str] = [
                    obj_record["name"] if i_line_index == 0 else "",
                    obj_record["car_no_display"] if i_line_index == 0 else "",
                    obj_record["car_no"] if i_line_index == 0 else "",
                    obj_record["car_no"] if i_line_index == 0 else "",
                    list_slot_values[0],
                    list_slot_values[1],
                    list_slot_values[2],
                    list_slot_values[3],
                    list_slot_values[4],
                    list_slot_values[5],
                    obj_record["title_text"] if i_line_index == 0 else "",
                    obj_record["body_text"].replace("\n", "\\n") if i_line_index == 0 else "",
                    obj_record["work_date_text"] if i_line_index == 0 else "",
                ]
                obj_tsv_file.write("\t".join(list_row_values) + "\n")

    with obj_step0002_json_path.open(mode="w", encoding="utf-8", newline="\n") as obj_json_file:
        for obj_record in list_records:
            obj_json_file.write(json.dumps(obj_record, ensure_ascii=False) + "\n")

    return str(obj_step0002_tsv_path), str(obj_step0002_json_path)


def convert_excel_to_tsv(psz_excel_file_path: str) -> str:
    """Convert active sheet of an Excel file to UTF-8 TSV with CRLF line endings."""
    obj_excel_path: Path = Path(psz_excel_file_path)
    psz_tsv_file_name: str = f"{obj_excel_path.stem}.tsv"
    obj_tsv_file_path: Path = obj_excel_path.with_name(psz_tsv_file_name)

    obj_workbook: Any = load_workbook(filename=str(obj_excel_path), data_only=True)
    obj_active_sheet: Any = obj_workbook.active

    list_skip_keywords: list[str] = ["始業前点検"]

    list_source_rows: list[list[str]] = []
    for obj_row in obj_active_sheet.iter_rows(values_only=True):
        list_row_values: list[str] = []
        for obj_cell_value in obj_row:
            if obj_cell_value is None:
                psz_cell_text: str = ""
            else:
                psz_cell_text = escape_newlines_in_cell_text(str(obj_cell_value))
            list_row_values.append(psz_cell_text)

        if should_skip_row_by_first_column(list_row_values, list_skip_keywords):
            continue

        list_source_rows.append(list_row_values)

    list_normalized_rows: list[list[str]] = merge_continuation_rows(list_source_rows)

    list_expanded_rows: list[list[str]] = expand_rows_by_embedded_newlines(list_normalized_rows)

    with obj_tsv_file_path.open(mode="w", encoding="utf-8", newline="\r\n") as obj_tsv_file:
        for list_normalized_row in list_expanded_rows:
            psz_tsv_line: str = "\t".join(list_normalized_row)
            obj_tsv_file.write(psz_tsv_line + "\n")

    obj_workbook.close()
    return str(obj_tsv_file_path)


def main() -> int:
    list_arguments: list[str] = sys.argv
    if len(list_arguments) < 2:
        psz_usage_message: str = "Usage: python DispatchCalendar_Cmd.py <excel_file_path1> [excel_file_path2 ...]"
        print(psz_usage_message)
        return 1

    list_excel_file_paths: list[str] = list_arguments[1:]
    i_success_count: int = 0
    i_failure_count: int = 0

    for psz_excel_file_path in list_excel_file_paths:
        obj_excel_path: Path = Path(psz_excel_file_path)

        if not obj_excel_path.exists() or not obj_excel_path.is_file():
            psz_error_message: str = f"Input file not found: {psz_excel_file_path}"
            print(psz_error_message)
            write_error_text(psz_excel_file_path, psz_error_message)
            i_failure_count += 1
            continue

        if obj_excel_path.suffix.lower() != ".xlsx":
            psz_error_message = f"Only .xlsx files are supported: {psz_excel_file_path}"
            print(psz_error_message)
            write_error_text(psz_excel_file_path, psz_error_message)
            i_failure_count += 1
            continue

        try:
            psz_created_tsv_path: str = convert_excel_to_tsv(psz_excel_file_path)
            print(f"TSV created: {psz_created_tsv_path}")
            psz_created_step_tsv_path: str = create_step0001_tsv_from_tsv(psz_created_tsv_path)
            print(f"Step TSV created: {psz_created_step_tsv_path}")
            psz_step0002_tsv_path, psz_step0002_json_path = create_step0002_outputs_from_step0001_tsv(psz_created_step_tsv_path)
            print(f"Step0002 TSV created: {psz_step0002_tsv_path}")
            print(f"Step0002 JSON created: {psz_step0002_json_path}")
            i_success_count += 1
        except Exception as obj_exception:  # noqa: BLE001
            psz_traceback_text: str = traceback.format_exc()
            psz_error_message = (
                "Failed to convert Excel to TSV.\n"
                f"Input: {psz_excel_file_path}\n"
                f"Error: {obj_exception}\n\n"
                f"Traceback:\n{psz_traceback_text}"
            )
            print(psz_error_message)
            write_error_text(psz_excel_file_path, psz_error_message)
            i_failure_count += 1

    print(f"Summary: success={i_success_count}, failure={i_failure_count}")
    if i_failure_count == 0:
        return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
